#depop start
import http.client
import json
import pandas as pd

# Set up connection to Depop API
conn = http.client.HTTPSConnection("depop-thrift.p.rapidapi.com")

# Set up headers
headers = {
    'x-rapidapi-key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'x-rapidapi-host': "depop-thrift.p.rapidapi.com"
}

# Make the request to Depop API
conn.request("GET", "/getSearch?page=100&keyword=nike%20mf%20doom&countryCode=us&sortBy=newlyListed", headers=headers)

# Get the response
res = conn.getresponse()
data = res.read()

#print(data.decode("utf-8"))

# Decode response data
response_str = data.decode("utf-8")

# Parse JSON response
response_json = json.loads(response_str)

# Print the JSON response to understand its structure
print(json.dumps(response_json, indent=4))

# Since the response is a list, use it directly
if isinstance(response_json, list):
    items = response_json
else:
    print("Unexpected JSON structure:", type(response_json))
    items = []
# Assuming the JSON response is a dictionary with a key 'results' containing a list of products
# Adjust the key according to the actual structure of the response
#items = response_json.get('results', [])

if not items:
    print("No items found in the JSON response.")

# Normalize JSON data to flat table
df = pd.json_normalize(items)

# Write DataFrame to CSV
df.to_csv('outputdepop.csv', index=False)

print("CSV file created successfully!")

#poshmark start
import http.client
import json
from openpyxl import Workbook
import re

# Function to clean the cell value
def clean_cell_value(value):
    if isinstance(value, str):
        # Remove non-printable characters
        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    return value

# Establish a connection to the Poshmark API
conn = http.client.HTTPSConnection("poshmark.p.rapidapi.com")

# Set the headers for the API request
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "poshmark.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/search?query={formatted_query}&domain=com", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    # Print the JSON response to understand its structure (for debugging)
    print(json.dumps(decoded_data, indent=4))

    # Adjust the following key to match the actual structure of the response
    items = decoded_data.get('data', [])

    if items:
        # Assuming each item in 'data' is a dictionary
        response_headers = items[0].keys() if items else []
        ws.append(list(response_headers))
        for item in items:
            row = []
            for header in response_headers:
                cell_value = item.get(header, '')
                # Convert complex structures to strings
                if isinstance(cell_value, (list, dict)):
                    cell_value = json.dumps(cell_value)
                # Clean the cell value
                cell_value = clean_cell_value(cell_value)
                row.append(cell_value)
            ws.append(row)
    else:
        print(f"No items found for query: {query}")
    time.sleep(0.5)

# Save the workbook to a file
wb.save("outputposhmark.xlsx")
print("Search results have been saved to outputposhmark.xlsx.")

#mercari
import http.client
import json
from openpyxl import Workbook

# Establish a connection to the API
conn = http.client.HTTPSConnection("mercari.p.rapidapi.com")
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "mercari.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/Mercari/Search?page=1&query={formatted_query}", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    if isinstance(decoded_data, list) and decoded_data:
        response_headers = decoded_data[0].keys() if decoded_data else []
        ws.append(list(response_headers))
        for item in decoded_data:
            row = []
            for header in response_headers:
                cell_value = item.get(header, '')
                # Convert complex structures to strings
                if isinstance(cell_value, (list, dict)):
                    cell_value = json.dumps(cell_value)
                row.append(cell_value)
            ws.append(row)

# Save the workbook to a file
wb.save("outputmercari.xlsx")
print("Search results have been saved to outputmercari.xlsx.")

#ebay start
import http.client
import json
import requests

url = "https://api.ebay.com/buy/browse/v1/item_summary/search?q=Nike%20SB%20MF%20DOOM&limit=10"

payload = {}
headers = {
  'Authorization': 'Bearer v^1.1#i^1#p^1#I^3#f^0#r^0#t^H4sIAAAAAAAAAOVYbWwURRi+6wflo0BiUUhBcm7hD83tze7tfeymd+F6beEEei13xZZEYG5vtl26t7vZnW17GmM9TdXEhChSRGNSiPwgQkiqNTHlw6gR8YeAgDGgRkBRERH5gYRf7l6Pcq0EkF7iJd6fy7zzzjvP88z7zswO6J82Y9nAyoG/ZtsrSob6QX+J3U7NAjOmldfOKS2pLreBPAf7UP+S/rJM6S91OkxJKrcW6aoi68jRl5JkncsaA4ShyZwCdVHnZJhCOod5LhZas5qjScCpmoIVXpEIR6QhQLgZgQW0n+cFD81CJJhW+VbMuBIg/ElIscDngdDD814/bfbruoEiso6hjAMEDWjGCbxOiolTfo5hOA8gvTSznnCsQ5ouKrLpQgIimIXLZcdqeVjvDhXqOtKwGYQIRkJNsWgo0tDYHK9z5cUK5nSIYYgNfWIrrCSRYx2UDHT3afSsNxczeB7pOuEKjs0wMSgXugXmAeBnpfa6E5BnaVZAwO9BLFsQKZsULQXx3XFYFjHpFLKuHJKxiNP3UtRUI7EZ8TjXajZDRBoc1l+rASVREJEWIBrrQx2hlhYiGIN8lyjHDGe0Lb4qEo85W9Y2OAHkgYfxUj4nTQsekzbMzTMWLKfypInCipwULc10R7OC65EJGk2Wxp0njekUlaNaSMAWoDw/GoxLCNZbazq2iAbukq1lRSlTB0e2ee8FGB+NsSYmDIzGI0zuyCoUIKCqiklicmc2FXPZ06cHiC6MVc7l6u3tJXvdpKJ1umgAKFf7mtUxvgulIGH5WrWe9RfvPcApZqnwyBypixxOqyaWPjNVTQByJxFkWDfDUjndJ8IKTrb+w5DH2TWxIApVIDzDu91ulqJo3utLMolCFEgwl6MuCwdKwLQzBbVuhFUJ8sjJm3lmpJAmJjm3R6DdfgE5k15WcDKsIDgTnqQ5mYAQQCiR4Fn//6hO7jfTY4jXEC5MqhcqzeP1fX460gOM1uZwNFwrrXX76N42uiG1goo11kZq1dWrxKhrcxQ1dQTutxjuSD4siaYycXP+4qv1lYqOUXJK9GK8oqIWRRL5dHEtsFtLtkANp2NIkkzDlEiGVDVSoK26UPT+3S7xYLQLeEL9N6fTHVnpVsYWFytrvG4GgKpIWucPySsplwINq9Zxl2XemEU9Jd6ieW0tKtYmyTG2YnLsvkmalHEXqffwpIZ0xdDMqzYZte5fcaUbyeZxhjVFkpC2bmoZYJVzKmVgmJBQsdV1ARJchEV21lI+yu9mgZcFU+LFZ0/SjcW2JRVoJy4LP8Cd2jXxAz9oy/6ojP1jkLEfKrHbQR1YStWAx6aVtpWVVlbrIkakCAVSFztl87tVQ2Q3SqtQ1EqqbNd2bVsZrm6MDi57Op4+/tYRW2Xe+8LQk2DB+AvDjFJqVt5zA1h0u6ecmjt/Ns0AL8VQfobxgPWg5nZvGfVI2bxhx/KF0w+U/HzYdmno9cBvtl/PLmgFs8ed7PZyW1nGbkNVp2t+2nEs8111xcLvex7/5MrpA+VnIiUXbhzfg778ZlPjjRVfO5n31WuLh8TheaeeG8yE39yx/2Dn/ksfHlwwZ0CbO/2jfc8fpdrY+ncXrhB8Gzq2n9g5+ujD2164efHF61UV3R9Ubh/dPVxvvzywdeZ7r54Ltu7tE4+7ftiz65ArdEw7u/TmyMjVy1dCI29cvbLhoY4wPvrMj5WbThze8if0oS1bbmQWLfm9cZju/GLOO2w72X5ycOZnf5QL/qea3t7Ebb1w7JVt148c+WoU9J6queqV1JG9y3f6UPW89EXx84oLL/ccPrevaevO668pZ156Nmwnnhisql4jtS+++e3umejT+eeN0bqT58fW8m8dqVvC+REAAA=='
}

response = requests.request("GET", url, headers=headers, data=payload)

print(response.json())

#
# conn = http.client.HTTPSConnection("api.ebay.com")
#
# headers = {
#     'Authorization': 'Bearer v^1.1#i^1#p^1#r^0#I^3#f^0#t^H4sIAAAAAAAAAOVYe2wURRi/64tAKVXwUashxxYkcNm72b2913p3yfUBPfu4wl1BiEj2MdsuvdtddmZpLxI5S0SDwYiGEoINSARCNAbU+CKa2GBIJEH5A4iaaIgoSogkGoMYEXe3pVwrAaSX2MT75zLffPPN7/eb75uZHZCvmLpwU/OmS1XOKSW78yBf4nRSlWBqRbl7RmlJbbkDFDg4d+fn5sv6S3+MIC6b0dilEGmqgqCrL5tREGsbo4ShK6zKIRmxCpeFiMUCm4q3tbK0B7CarmJVUDOEK9EYJcQQxcMgJ0g+fyAI+aBpVa7FTKtRgg/xDEMBIPKSJArhkNmPkAETCsKcgqMEDWiGBH6SYtIAsD4fywQ8IR+1knAtgzqSVcV08QAiZsNl7bF6AdabQ+UQgjo2gxCxRHxRKhlPNDa1pyPeglixER1SmMMGGttqUEXoWsZlDHjzaZDtzaYMQYAIEd7Y8Axjg7Lxa2DuAL4tdRjAoJ/iKT4AeTokgKJIuUjVsxy+OQ7LIoukZLuyUMEyzt1KUVMNfg0U8Eir3QyRaHRZf0sMLiNLMtSjRFN9fEW8o4OIpTihW1ZSBpnsTLck0imyY2kjCTgB+JkAFSRpWvKHJciNzDMcbETlcRM1qIooW5ohV7uK66EJGo6Xhi6QxnRKKkk9LmELUKEffU1COrjSWtPhRTRwt2ItK8yaOrjs5q0XYHQ0xrrMGxiORhjfYSsUJThNk0VifKediiPZ04eiRDfGGuv19vb2enp9HlXv8tIAUN7H2lpTQjfMcoTla9W67S/fegAp21QEaI5EMotzmomlz0xVE4DSRcSYsI8JUyO6j4UVG2/9h6GAs3dsQRSrQPggHWZCYVFkOBD2cVIxCiQ2kqNeCwfkuRyZ5fQeiLUMJ0BSMPPMyEJdFlmfX6J9IQmSYiAskUxYkkjeLwZISoIQQMjz5ub3P6qT2830FBR0iIuT6sVK83R9X4hOrAPGkvaGZIM7s9QXpHs76cbsYirV5E64tdYWOeldk4SLVkRvtxhuSL4hI5vKpM35J1+tN6sIQ3FC9FKCqsEONSMLucm1wD5d7OB0nEvBTMY0TIhkXNMSRdqqi0Xv3+0Sd0a7iCfUf3M63ZAVsjJ2crGyxiMzAKfJHuv88Qhq1qtyhlXruNsyr7ZRT4i3bF5bJxVrk+QwW1kcvm96TMq424PWCR4dItXQzau2J2ndv9JqD1TM4wzraiYD9WUTywCrnLNZA3N8Bk62ui5CgsvcJDtrqSDlDzCBEB2YEC/BPklXT7YtqUg7cVnDHdypvWM/8GMO+0f1O4dAv/PjEqcTRMA8qg7MqSjtLCudXotkDD0yJ3mQ3KWY36069PTAnMbJesksxy97tjU31DYlBxY+mc59sfOoY3rB+8LuVaBm9IVhailVWfDcAB663lNOVd9fRTPATzEA+HxMYCWou95bRt1Xds+BvWy120sdrnk71/fOsZqfHY70EKgadXI6yx1l/U5HpLNq3Zw9S57eVtd18tQ0bduOh4fIjZfP7/0r8e3JQxsSy33rj5e/eejsa7MHT/yW2vk85V770ov7B77evuXVeCQ545lpiYvVX66vWrvjaPypWPCBC1sXv/XNqg/PL+g69P7MrUZL+N4z4iOb8/UDwgcLd7wS8b+g7K/E80vOX/xkzq/bNx+csjPfeeHTwbl3z74yvwdVxdoi7VWX93m/Wv7dTHnrrBMnHnTMOxaoOdv2Obdv3uHT6y+fO8cMvOz+vWnGu6f2/dmyR/ro0TObYq18qOPK1Z+OVG85+/2xXbUbn43Pr7v6xHs/sH907bp0oHn16+7+iv2BIwePv7G2fXDDc6fdlbHawbtmD4XQgsc/G15Kx9+6FGG7+REAAA=='
# }
#
# conn.request("GET", "/buy/browse/v1/item_summary/search?q=nike&limit=10", headers=headers)
#
# res = conn.getresponse()
# data = res.read()
#
# # Decode the bytes data to a string
# data_str = data.decode("utf-8")
# print(data_str)

if response.status_code == 200:
    # Extract JSON content from the response
    json_content = response.json()

    # Save the JSON content to a file
    with open('ebaydata.json', 'w') as f:
        json.dump(json_content, f)

    print("JSON file saved successfully.")
else:
    print("Error:", response.status_code)

import pandas as pd
from pandas import json_normalize
import json

json_path = "ebaydata.json"

with open(json_path, "r") as file:
    json_data = json.load(file)

if 'itemSummaries' in json_data:
    item_summaries = json_data['itemSummaries']
    print(item_summaries)
else:
    print("Key 'itemSummaries' not found in the JSON data.")

df = json_normalize(item_summaries)

df.to_csv('ebaydata.csv', index=False)
