#OUTKITS ALL - 07/01/24 - email code at the bottom, each excel file has uniform column names and column quantity.
#Single excel files for each API -RF 07/01/24

import http.client
import json
import time
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
import re
import requests
from pandas import json_normalize
from urllib.parse import quote
#import src.ebayconnection


# Function to clean the cell value
def clean_cell_value(value):
    if isinstance(value, str):
        # Remove non-printable characters
        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    return value

# Function to create a keyword query for multiple search terms
def create_keyword_query(keywords):
    return quote(" ".join(keywords))

# Establish a connection to the Depop API
conn = http.client.HTTPSConnection("depop-thrift.p.rapidapi.com")

# Set up headers
headers = {
    'x-rapidapi-key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'x-rapidapi-host': "depop-thrift.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk MF DOOM",
    "Nike Dunk SB Stussy",
    "Nike Dunk Pushead",
    "Nike Dunk High SB Khaki Creed",
    "Nike Dunk 6.0 Hemp",
    "Nike Dunk SB Mocha",
    "Nike Dunk SB Bison",
    "Nike Dunk SB Mocha Choc",
    "Nike Dunk SB Medusa",
    "Nike Dunk SB Oompa Loompa",
    "Nike 6.0 NKE Quasar Purple",
    "Nike Dunk SB Crown Royal",
    "Nike Dunk Palm Green",
    "Nike Dunk Low Cargo Khaki",
    "Nike Dunk CL denim",
    "Nike Dunk Low Pro Mushroom",
    "Nike Dunk Low SB Tweed",
    "Nike Dunk Low ACG",
    "Dunk Low Pro Obsidian",
    "Dunk Low Pro Midnight Navy",
    "Nike Dunk Low Pro Mesa",
    "Dunk Low Pro B Olive",
    "Nike SB Dunk Trail End",
    "Nike SB Dunk Dusty Cactus",
    "Pro B Oxide"
]

# Create a new Excel workbook and a single sheet
wb = Workbook()
ws = wb.active
ws.title = "Consolidated Results"

# Define the desired columns with the new headers and order
desired_columns = {
    'slug': 'Name',
    'price': 'Price',
    'sizes': 'Size',
    'url': 'URL'
}
# New header order
header_order = ['Name', 'Price', 'Size', 'Gender', 'URL']

# Append the headers to the sheet
ws.append(header_order)

# Process each query
for query in queries:
    keywords = query.split(" ")
    keyword_query = create_keyword_query(keywords)
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Initialize variable to track if items were found and retry count
    items_found = False
    retries = 0
    max_retries = 10

    # Retry until items are found or max retries reached
    while not items_found and retries < max_retries:
        # Make the GET request with the correct headers
        conn.request("GET", f"/getSearch?page=1&keyword={keyword_query}&countryCode=us&sortBy=newlyListed", headers=headers)
        res = conn.getresponse()
        data = res.read()

        time.sleep(2.0)

        # Decode the JSON response
        decoded_data = json.loads(data.decode("utf-8"))

        # Print the JSON response to understand its structure (for debugging)
        print(json.dumps(decoded_data, indent=4))

        # Check if decoded_data is a list
        if isinstance(decoded_data, list):
            items = decoded_data
        else:
            # Adjust the following key to match the actual structure of the response
            items = decoded_data.get('products', [])

        # Filter items to include only those with the brand 'Nike'
        nike_items = [item for item in items if item.get('brand') and item.get('brand').lower() == 'nike']

        if nike_items:
            items_found = True
            print(f"Items have been found for query: {query}")
            for item in nike_items:
                row = []
                for header in header_order:
                    if header == 'Gender':
                        cell_value = ''  # Blank column for Gender
                    else:
                        cell_key = list(desired_columns.keys())[list(desired_columns.values()).index(header)]
                        if cell_key == 'price':
                            price_info = item.get('price', {})
                            amount = float(price_info.get('amount', 0))
                            national_shipping = float(price_info.get('nationalShipping', 0))
                            cell_value = amount + national_shipping
                        else:
                            cell_value = item.get(cell_key, '')
                        if isinstance(cell_value, (list, dict)):
                            cell_value = json.dumps(cell_value)
                        cell_value = clean_cell_value(cell_value)
                    row.append(cell_value)
                ws.append(row)
        else:
            retries += 1
            print(f"No items found for query: {query}, retrying... (Attempt {retries}/{max_retries})")

    if not items_found:
        print(f"Max retries reached for query: {query}. Moving on to the next keyword.")

timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

# Save the workbook to a file with a timestamp
filename = f"outputdepop_{timestamp}.xlsx"
wb.save(filename)
print(f"Search results have been saved to {filename}.")

#poshmark start


# Function to clean the cell value
def clean_cell_value(value):
    if isinstance(value, str):
        # Remove non-printable characters
        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    return value

# Establish a connection to the Poshmark API
conn = http.client.HTTPSConnection("poshmark.p.rapidapi.com")

# Set the headers for the API request
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "poshmark.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/search?query={formatted_query}&domain=com", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    # Print the JSON response to understand its structure (for debugging)
    print(json.dumps(decoded_data, indent=4))

    # Adjust the following key to match the actual structure of the response
    items = decoded_data.get('data', [])

    if items:
        # Assuming each item in 'data' is a dictionary
        response_headers = items[0].keys() if items else []
        ws.append(list(response_headers))
        for item in items:
            row = []
            for header in response_headers:
                cell_value = item.get(header, '')
                # Convert complex structures to strings
                if isinstance(cell_value, (list, dict)):
                    cell_value = json.dumps(cell_value)
                # Clean the cell value
                cell_value = clean_cell_value(cell_value)
                row.append(cell_value)
            ws.append(row)
    else:
        print(f"No items found for query: {query}")
    time.sleep(2.0)

# Save the workbook to a file
wb.save("outputposhmark.xlsx")
print("Search results have been saved to outputposhmark.xlsx.")

#mercari

# Establish a connection to the API
conn = http.client.HTTPSConnection("mercari.p.rapidapi.com")
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "mercari.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/Mercari/Search?page=1&query={formatted_query}", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    if isinstance(decoded_data, list) and decoded_data:
        response_headers = decoded_data[0].keys() if decoded_data else []
        ws.append(list(response_headers))
        for item in decoded_data:
            row = []
            for header in response_headers:
                cell_value = item.get(header, '')
                # Convert complex structures to strings
                if isinstance(cell_value, (list, dict)):
                    cell_value = json.dumps(cell_value)
                row.append(cell_value)
            ws.append(row)

# Save the workbook to a file
wb.save("outputmercari.xlsx")
print("Search results have been saved to outputmercari.xlsx.")

#ebay start

#connection = src.EbayConnection()
#connection.connect()

