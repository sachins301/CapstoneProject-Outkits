# AR - Feel free too add the ebay code after mercari code and the outputebay.xlsx file in (attachment_paths) list at the end so it got included. 
# My computer is not runnnign the ebay only. 
import http.client
import json
import time
import pandas as pd
from openpyxl import Workbook
import re
import requests
from pandas import json_normalize
from urllib.parse import quote
import importlib.util
import os
import smtplib
from email.message import EmailMessage

#depop start
# Function to clean the cell value
def clean_cell_value(value):
    if isinstance(value, str):
        # Remove non-printable characters
        value = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', value)
    return value

# Function to create a keyword query for multiple search terms
def create_keyword_query(keywords):
    return quote(" ".join(keywords))

# Establish a connection to the Depop API
conn = http.client.HTTPSConnection("depop-thrift.p.rapidapi.com")

# Set up headers
headers = {
    'x-rapidapi-key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'x-rapidapi-host': "depop-thrift.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk MF DOOM",
    "Nike Dunk SB Stussy",
    "Nike Dunk Pushead",
    "Nike Dunk High SB Khaki Creed",
    "Nike Dunk 6.0 Hemp",
    "Nike Dunk SB Mocha",
    "Nike Dunk SB Bison",
    "Nike Dunk SB Mocha Choc",
    "Nike Dunk SB Medusa",
    "Nike Dunk SB Oompa Loompa",
    "Nike 6.0 NKE Quasar Purple",
    "Nike Dunk SB Crown Royal",
    "Nike Dunk Palm Green",
    "Nike Dunk Low Cargo Khaki",
    "Nike Dunk CL demim",
    "Nike Dunk Low Pro Mushroom",
    "Nike Dunk Low SB Tweed",
    "Nike Dunk Low ACG",
    "Dunk Low Pro Obsidian",
    "Dunk Low Pro Midnight Navy",
    "Nike Dunk Low Pro Mesa",
    "Dunk Low Pro B Olive",
    "Nike SB Dunk Trail End",
    "Nike SB Dunk Dusty Cactus",
    "Pro B Oxide"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    keywords = query.split(" ")
    keyword_query = create_keyword_query(keywords)
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/getSearch?page=100&keyword={keyword_query}&countryCode=us&sortBy=newlyListed", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    # Print the JSON response to understand its structure (for debugging)
    #print(json.dumps(decoded_data, indent=4))

    # Check if decoded_data is a list
    if isinstance(decoded_data, list):
        items = decoded_data
    else:
        # Adjust the following key to match the actual structure of the response
        items = decoded_data.get('products', [])

    if items:
        # Assuming each item in 'products' is a dictionary
        response_headers = items[0].keys() if items else []
        ws.append(list(response_headers))
        for item in items:
            row = []
            for header in response_headers:
                cell_value = item.get(header, '')
                # Convert complex structures to strings
                if isinstance(cell_value, (list, dict)):
                    cell_value = json.dumps(cell_value)
                # Clean the cell value
                cell_value = clean_cell_value(cell_value)
                row.append(cell_value)
            ws.append(row)
    else:
        print(f"No items found for query: {query}")
    time.sleep(1.0)

# Save the workbook to a file
wb.save("outputdepop.xlsx")
print("Search results have been saved to outputdepop.xlsx.")


#Establish a connection to the Poshmark API

# Function to extract the number from a string and remove the last zero (only for Poshmark)
def extract_number(value):
    number = ''.join(filter(str.isdigit, value))
    if number.endswith('0'):
        number = number[:-1]
    return number


conn = http.client.HTTPSConnection("poshmark.p.rapidapi.com")

# Set the headers for the API request
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "poshmark.p.rapidapi.com"
}

# Verify that headers are correct (for debugging)
print("Headers type before request:", type(headers))

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/search?query={formatted_query}&domain=com", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    # Print the JSON response to understand its structure (for debugging)
    #print(json.dumps(decoded_data, indent=4))

    # Adjust the following key to match the actual structure of the response
    items = decoded_data.get('data', [])

    if items:
        # Define the headers for the columns we are interested in
        ws.append(['Name', 'Price', 'Size', 'Gender', 'URL'])

        for item in items:
            title = item.get('title', '')
            price = extract_number(item.get('price_amount', {}).get('val', ''))
            size_obj = item.get('size_obj', {}).get('display', '')

            # Handle cases where size_obj might be an empty dictionary or None
            size = size_obj if size_obj else ''

            department = item.get('department', {}).get('display', '')

            # Construct the URL
            clean_title = title.replace(' ', '-').replace('/', '-')
            item_id = item.get('id', '')
            url = f"https://poshmark.com/listing/{clean_title}-{item_id}"

            row = [title, price, size, department, url]
            row = [clean_cell_value(cell) for cell in row]
            ws.append(row)
    else:
        print(f"No items found for query: {query}")
    time.sleep(0.5)

# Save the workbook to a file
wb.save("outputposhmark.xlsx")
print("Search results have been saved to outputposhmark.xlsx.")



# Establish a connection to the MERCARI API
conn = http.client.HTTPSConnection("mercari.p.rapidapi.com")
headers = {
    'X-RapidAPI-Key': "e68937903dmsh53a9dc44eeefef1p1b6bedjsn8e1076d22514",
    'X-RapidAPI-Host': "mercari.p.rapidapi.com"
}

# List of queries to process
queries = [
    "Nike Dunk Low 6.0",
    "Nike Dunk Pro B",
    "Nike Dunk Pushead",
    "Nike SB Dunk Low Bigfoot",
    "Carhartt Hooded Jacket Red",
    "Carhartt Hooded Jacket Orange",
    "Carhartt Hooded Jacket Teal",
    "Carhartt Hooded Jacket Purple",
    "Nike SB Dunk Size 12",
    "Nike Dunk Cinder Brown",
    "Nike Dunk Light Stone Bone"
]

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each query
for query in queries:
    formatted_query = query.replace(" ", "%20")
    print(f"Requesting data for query: {query}")  # Debugging statement

    # Make the GET request with the correct headers
    conn.request("GET", f"/Mercari/Search?page=1&query={formatted_query}", headers=headers)
    res = conn.getresponse()
    data = res.read()

    # Decode the JSON response
    decoded_data = json.loads(data.decode("utf-8"))

    # Create a new sheet for each query
    ws = wb.create_sheet(title=query[:31])  # Excel sheet names are limited to 31 characters

    if isinstance(decoded_data, list) and decoded_data:
        response_headers = ['Name (listing title)', 'Price', 'Size', 'Gender', 'URL']
        ws.append(response_headers)
        for item in decoded_data:
            name = item.get('name', '')
            price = item.get('price', '')
            size_dict = item.get('itemSize', {})
            size = size_dict.get('name', '') if isinstance(size_dict, dict) else ''
            size = size.split(' ')[0].replace(',', '').replace('(', '').replace(')', '').replace(" ", '0')
            gender = item.get('categoryTitle', '')
            url = item.get('url', '')
            row = [name, price, size, gender, url]
            ws.append(row)

# Save the workbook to a file
wb.save("outputmercari.xlsx")
print("Search results have been saved to outputmercari.xlsx.")


# Function to send email - Make sure to have the xslx files ready because this would be attached.
def send_email(subject, body, to, attachment_paths):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = "outkitsteam@gmail.com"
    msg['To'] = to
    msg.set_content(body)

    # Add the attachments
    for attachment_path in attachment_paths:
        if os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(attachment_path)
            msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
        else:
            print(f"File {attachment_path} does not exist.")

    email_password = 'okui erov ucsn drev'  # Replace with the generated app-specific password

    # Send the email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login("outkitsteam@gmail.com", email_password)
            smtp.send_message(msg)
        print("Email alert has been sent.")
    except smtplib.SMTPAuthenticationError as e:
        print(f"SMTP Authentication Error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

# # Main execution
# # fetch_depop_data()
# fetch_poshmark_data()
# fetch_mercari_data()

attachment_paths = ['outputmercari.xlsx', 'outputposhmark.xlsx', 'outputdepop.xlsx']
send_email(
    subject="Search Results for APIS",
    body="Please find attached the search results.",
    to="u1482560@utah.edu",
    attachment_paths=attachment_paths
)

print('Successfully sent the mail')

